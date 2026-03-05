import logging
import os
import re
import json
from datetime import date
from openpyxl import load_workbook
from pathlib import Path
import argparse
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage

load_dotenv()


class Chunk:

    def __init__(self, chunk_id: str, chunk_text: str):
        self.chunk_id = chunk_id
        self.chunk_text = chunk_text
        self.parent_id = ""
        self.children = []
        self.subpoint = False

    def add_child(self, chunk_id: str):
        self.children.append(chunk_id)

    def add_parent(self, chunk_id: str):
        self.parent_id = chunk_id


class ChunkTree:

    def __init__(self):
        self.chunks = {}

    def add_chunk(self, chunk_id: str, chunk_text: str):
        new_chunk = Chunk(chunk_id, chunk_text)
        self.chunks[chunk_id] = new_chunk

    def add_edge(self, parent_id: str, child_id: str):
        self.chunks[parent_id].add_child(child_id)
        self.chunks[child_id].add_parent(parent_id)

    def get_children(self, chunk_id: str):
        return self.chunks[chunk_id].children


def numeric_sort_key(chunk_id: str) -> tuple:
    base = re.sub(r"[а-яА-Яa-zA-Z]\)$", "", chunk_id)
    parts = base.split(".")
    return tuple(int(p) for p in parts if p.isdigit())

# Функиция загружает словарь аббревиатур из JSON-файла
def load_abbreviations(filepath: str, logger) -> dict:
    with open(filepath, "r", encoding="utf-8") as f:
        abbreviations = json.load(f)
    logger.info(msg=f"Abbreviations loaded: {len(abbreviations)} entries from {filepath}")
    return abbreviations


# Функция загружает глоссарий терминологии из JSON и строит поисковый индекс
# Словарь terminology хранит пары term_key → запись (full_name, aliases, definition, distinguish_from, abbreviation, search_roots)
# search_index хранит пары full_name : term_key и abbreviation : term_key
def load_terminology(filepath: str) -> tuple[dict, dict]:
    with open(filepath, "r", encoding="utf-8") as f:
        terminology = json.load(f)
    search_index = {}
    for term_key, entry in terminology.items():
        search_index[entry["full_name"].lower()] = term_key
        if entry["abbreviation"]:
            search_index[entry["abbreviation"].lower()] = term_key
    return terminology, search_index

# Функция проверяет, содержит ли текст хотя бы одну аббревиатуру из словаря
# Используется для заполнения метаданных has_abbreviations
def detect_abbreviations(text: str, abbreviations: dict) -> bool:
    for abbr in abbreviations:
        if re.search(r"(?<!\w)" + re.escape(abbr) + r"(?!\w)", text):
            return True
    return False


# Функция ищет в тексте аббревиатуры из словаря аббревиатур
def find_abbreviations_in_text(text: str, abbreviations: dict) -> dict:
    found = {}
    for abbr, definition in abbreviations.items():
        is_latin = bool(re.match(r"^[a-zA-Z]", abbr))
        if is_latin:
            pattern = re.compile(r"(?<!\w)" + re.escape(abbr) + r"(?!\w)", re.IGNORECASE)
        else:
            pattern = re.compile(r"(?<!\w)" + re.escape(abbr) + r"(?!\w)")
        if pattern.search(text):
            found[abbr] = definition
    return found


# Функция ищет в тексте все термины и аббревиатуры по заданной терминологии
def find_terms_in_text(text: str, search_index: dict,
                       terminology: dict) -> list[str]:
    text_lower = text.lower()
    found = set()

    # Путь 1 и 2: точный поиск по full_name и abbreviation
    for key, term_key in search_index.items():
        pattern = re.compile(
            r"(?<!\w)" + re.escape(key) + r"(?!\w)", re.IGNORECASE
        )
        if pattern.search(text_lower):
            found.add(term_key)

    # Путь 3: AND-поиск по search_roots
    # ВСЕ корни из списка должны присутствовать в тексте
    for term_key, entry in terminology.items():
        roots = entry.get("search_roots", [])
        if roots and all(r.lower() in text_lower for r in roots):
            found.add(term_key)

    return list(found)

# Функция формирует текстовый блок с определениями терминов для добавления в системный промпт LLM
# Возвращает пустую строку, если релевантных терминов не найдено
def build_glossary_block(found_terms: list[str], terminology: dict) -> str:
    if not found_terms:
        return ""

    lines = ""
    for term_key in found_terms:
        entry = terminology[term_key]
        name = entry["full_name"]
        if entry["abbreviation"]:
            name += f" ({entry['abbreviation']})"
        lines += f"\n- {name}: {entry['definition']}"

    return (
            "\n\nНекоторые термины в тексте требуют раскрытия. "
            "Если встречается термин из списка ниже — вплети его суть "
            "одним уточняющим предложением прямо в абзац, своими словами.\n\n"

            "Пример как делать ПРАВИЛЬНО:\n"
            "Исходный пункт: «Формирование плановых партий — "
            "группировка по толщине материала»\n"
            "Определение из словаря: «Плановая партия — расчётный объект MRP: "
            "несколько потребностей в одинаковых ДСЕ объединяются в одну партию»\n"
            "Результат: «При формировании плановых партий — расчётных объектов MRP, "
            "объединяющих потребности в одинаковых ДСЕ — система поддерживает "
            "группировку по технологическому критерию толщины материала.»\n\n"

            "Пример как делать НЕПРАВИЛЬНО:\n"
            "Результат: «При формировании плановых партий система поддерживает "
            "группировку по толщине материала. Плановая партия — это расчётный "
            "объект MRP, который объединяет потребности в одинаковых ДСЕ.»\n"
            "Почему плохо: определение вынесено отдельным предложением в конец, "
            "а не вплетено в контекст.\n\n"

            "Термины для раскрытия:"
            + lines
    )


# Функция заменяет аббревиатуры в тексте на их расшифровки
def expand_abbreviations(text: str, abbreviations: dict) -> str:
    # Сортируем по убыванию длины, чтобы длинные аббревиатуры искались раньше (MRP-II раньше MRP)
    for abbr in sorted(abbreviations, key=len, reverse=True):
        escaped = re.escape(abbr)
        # Пропускаем если аббревиатура уже расшифрована: паттерн "АББР (расшифровка)"
        already_expanded = re.compile(
            r"(?<!\w)" + escaped + r"\s*\([^)]*\)", re.IGNORECASE
        )
        if already_expanded.search(text):
            continue
        pattern = re.compile(r"(?<!\w)(" + escaped + r")(?!\w)", re.IGNORECASE)
        text = pattern.sub(r"\1 (" + abbreviations[abbr] + ")", text)
    return text


# Функция создаёт объект ChunkTree, добавляет узлы и рёбра
def create_chunk_tree(ws, logger):
    new_chunk_tree = ChunkTree()
    row_map = {}

    last_point = ""
    for row_id, i in enumerate(ws.values, start=1):
        if row_id == 1: # Пропускаем строку заголовков
            continue
        # Пропускаем пустые строки
        if not i or i[0] is None:
            continue

        point_id = str(i[0]).strip()

        # Проверяем, существует ли чанк с таким номером
        if point_id in new_chunk_tree.chunks.keys():
            logger.warning(msg=f"Duplicate point found: {point_id}")
        else:
            # Добавляем чанк если строка содержит пункт в правильном формате
            if re.match(r"\d+(\.\d+)*", point_id):

                new_chunk_tree.add_chunk(point_id, i[1])
                row_map[point_id] = row_id
                last_point = point_id

            # Добавляем чанк если строка содержит подпункт в правильном формате
            elif re.match(r"[а-яА-Яa-zA-Z]\)$", point_id):
                new_chunk_tree.add_chunk(last_point + point_id, i[1])
                row_map[last_point + point_id] = row_id

            else:
                logger.warning(msg=f"Point in wrong format found: {point_id}")

    for n in new_chunk_tree.chunks.keys():
        # Подпункты вида 4.6.11б) — убираем букву, чтобы получить id родителя
        if re.search(r"[а-яА-Яa-zA-Z]\)$", n):
            parent_id = re.sub(r"[а-яА-Яa-zA-Z]\)$", "", n)
            if parent_id in new_chunk_tree.chunks:
                new_chunk_tree.add_edge(parent_id, n)
        # Числовые подпункты вида 4.6.11 — убираем последний сегмент, чтобы получить id родителя
        elif "." in n:
            parent_id = n.rsplit(".", 1)[0]
            if parent_id in new_chunk_tree.chunks:
                new_chunk_tree.add_edge(parent_id, n)
    return new_chunk_tree, row_map


# Функция строит цепочку родителей и считает глубину уровня для любого узла
def build_parent_chain(logger, chunk_tree, chunk_id: str, parent_chain, level):
    # Проверяем, отсутствует ли родитель
    if chunk_tree.chunks[chunk_id].parent_id == "" and chunk_id.isdigit() == False:
        logger.warning(f"{chunk_id} is missing a parent")
        return level, parent_chain

    elif chunk_tree.chunks[chunk_id].parent_id == "":
        return level, parent_chain
    else:
        parent_chain.append(chunk_tree.chunks[chunk_id].parent_id)
        level += 1
        return build_parent_chain(logger, chunk_tree, chunk_tree.chunks[chunk_id].parent_id, parent_chain, level)


# Функция формирует текст для листового чанка с учётом до 3 ближайших родителей
def create_leaf_chunk_text(chunk_tree, chunk_id: str, parent_chain: list):
    parents = parent_chain[-3:] if len(parent_chain) > 3 else parent_chain
    chunk_text = ""
    space = "  "
    for id in parents:
        chunk_text += f"[{id}] {chunk_tree.chunks[id].chunk_text}\n{space}→  "
        space += "  "
    chunk_text += f"[{chunk_id}] {chunk_tree.chunks[chunk_id].chunk_text}"
    return chunk_text, len(parents)


# Функция формирует текст для группового чанка с учётом до 3 ближайших родителей
def create_group_chunk_text(chunk_tree, chunk_id: str, parent_chain, children):
    parents = parent_chain[-3:] if len(parent_chain) > 3 else parent_chain
    chunk_text = ""
    space = "  "
    for parent_id in parents:
        chunk_text += f"[{parent_id}] {chunk_tree.chunks[parent_id].chunk_text}\n{space}→  "
        space += "  "
    chunk_text += f"[{chunk_id}] {chunk_tree.chunks[chunk_id].chunk_text}\n    Включает подпункты:\n"
    for child_id in children:
        chunk_text += f"    - [{child_id}] {chunk_tree.chunks[child_id].chunk_text}\n"
    return chunk_text, len(parents)


# Функция формирует текст для векторизации чанка с помощью LLM
def create_chunk_text_context(chunk_tree, chunk_id: str, parent_chain: list, system_prompt,
                              terminology: dict, search_index: dict,
                              abbreviations: dict, plain_text: str):
    found_abbr = find_abbreviations_in_text(plain_text, abbreviations)
    found_terms = find_terms_in_text(plain_text, search_index, terminology)
    glossary_block = build_glossary_block(found_terms, terminology)

    extra = ""

    if found_abbr:
        extra += "\n---\nАббревиатуры встреченные в этом пункте:\n"
        for abbr, definition in found_abbr.items():
            extra += f"- {abbr}: {definition}\n"

    if glossary_block:
        extra += "\n---\nТермины встреченные в этом пункте:"
        extra += glossary_block
        extra += "\n---\nПри перефразировании раскрой эти термины согласно определениям выше прямо в тексте формулировки."

    if extra:
        augmented_prompt = SystemMessage(system_prompt.content + extra)
    else:
        augmented_prompt = system_prompt

    # Формируем текст чанка с учётом до 3 родителей для передачи в LLM
    parents = parent_chain[-3:] if len(parent_chain) > 3 else parent_chain
    chunk_text = ""
    space = "  "
    for pid in parents:
        chunk_text += f"[{pid}] {chunk_tree.chunks[pid].chunk_text}\n{space}→  "
        space += "  "
    chunk_text += f"[{chunk_id}] {chunk_tree.chunks[chunk_id].chunk_text}"

    user_prompt = HumanMessage(f"Преобразуй следующую иерархическую метку:{chunk_text}")
    messages = [augmented_prompt, user_prompt]
    response = llm.invoke(messages)

    return response.content, found_terms


# Функция сохраняет словарь в JSON-файл
def save_to_json(data, filepath):
    path = Path(filepath)

    # Создаём директорию если она не существует
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open(mode="w", encoding="utf-8") as f:
        json.dump(data, f, indent=1, ensure_ascii=False)


# Функция создаёт логгер
def create_logger():
    logger = logging.getLogger("chunk_function_bank_logger")
    logger.setLevel(logging.INFO)

    # Указываем путь к файлу логов
    log_path = Path("chunk_function_bank.log")
    log_path.parent.mkdir(parents=True, exist_ok=True)

    if not logger.handlers:
        file_handler = logging.FileHandler("chunk_function_bank.log")
        file_handler.setLevel(logging.INFO)

        formatter = logging.Formatter(
            "%(asctime)s | %(levelname)s | %(message)s"
        )

        file_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
    return logger


######################################################################################


# Создаём парсер аргументов командной строки
parser = argparse.ArgumentParser()
parser.add_argument("--input", required=True)
parser.add_argument("--terminology", default="dictionary/terminology_for_llm_full.json")
parser.add_argument("--abbreviations", default="dictionary/abbreviations.json")
args = parser.parse_args()

# Пути к входному и выходному файлам
input_path = args.input
today = date.today().strftime("%Y-%m-%d")
output_path = f"results/chunks_{today}.json"
terminology_path = args.terminology

# Создаём логгер
new_logger = create_logger()

new_logger.info(msg="Script started")

# Открываем файл и загружаем данные в рабочий лист openpyxl
try:
    file_path = Path(input_path)
    wb = load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]

    vec_col_id = None
    for cell in ws[1]:
        if cell.value == "Текст для векторизации":
            vec_col_id = cell.column
            break

    if vec_col_id is None:
        vec_col_id = ws.max_column + 1
        ws.cell(row=1, column=vec_col_id, value="Текст для векторизации")
        new_logger.info(f"'Текст для векторизации' column not found — created at column {vec_col_id}")

except Exception as e:
    new_logger.error(msg=e)
    print(e)
    exit()

# Создаём дерево чанков из Excel-файла
chunk_tree, row_map = create_chunk_tree(ws, new_logger)
new_logger.info(msg="Chunk tree created")

# Загружаем глоссарий терминологии и строим поисковый индекс
terminology, search_index = load_terminology(terminology_path)
new_logger.info(msg=f"Terminology loaded: {len(terminology)} terms, {len(search_index)} search keys")

# Загружаем словарь аббревиатур
abbreviations = load_abbreviations(args.abbreviations, new_logger)

# Создаём переменную для хранения результатов
result_data = []

# Создаём модель LLM
llm = ChatOpenAI(
    model=os.environ["OPENAI_MODEL"],
    api_key=os.environ["OPENAI_API_KEY"],
    temperature=0.0,
    seed=47
)

# Системный промпт
system_prompt = SystemMessage("""
Твоя задача — переформулировать иерархическую техническую метку
в единый связный абзац для последующей векторизации.

Правила:
- Раскрой все аббревиатуры прямо в тексте
- Сохрани ВСЕ конкретные детали из исходного текста —
  ни одна деталь не должна быть потеряна
- Не добавляй общих фраз ("позволяет оптимизировать",
  "повышает эффективность" и т.п.) — только то что есть в метке
- Включи контекст родительских пунктов в формулировку
- Пиши одним связным абзацем без заголовков, меток и списков
- Результат должен быть понятен без исходного текста
""")

total_chunks = len(chunk_tree.chunks)
print(f"\nProcessing {total_chunks} chunks...\n")

for idx, i in enumerate(chunk_tree.chunks.values(), 1):
    print(f"  [{idx}/{total_chunks}] {i.chunk_id} ({('group' if i.children else 'leaf')})", flush=True)
    result_chunk = dict()
    metadata = dict()

    level = 1
    parent_chain = []

    # Строим цепочку родителей и считаем глубину уровня
    level, parent_chain = build_parent_chain(new_logger, chunk_tree, i.chunk_id, parent_chain, level)
    parent_chain = sorted(parent_chain, key=numeric_sort_key)

    node_type = "group" if i.children else "leaf"
    metadata["chunk_id"] = "FB-" + i.chunk_id
    metadata["level"] = level
    metadata["node_type"] = node_type
    metadata["parent_id"] = "FB-" + i.parent_id if i.parent_id != "" else ""
    metadata["parent_chain"] = parent_chain
    metadata["children"] = i.children
    metadata["children_count"] = len(i.children)
    metadata["source_type"] = "function_bank"
    metadata["doc_name"] = Path(input_path).name

    result_chunk["chunk_id"] = "FB-" + i.chunk_id
    if node_type == "leaf":
        plain_text, parents_depth = create_leaf_chunk_text(chunk_tree, i.chunk_id, parent_chain)
    else:
        plain_text, parents_depth = create_group_chunk_text(chunk_tree, i.chunk_id, parent_chain, i.children)

    raw_plain_text = plain_text
    plain_text = expand_abbreviations(plain_text, abbreviations)

    # Получаем текст для векторизации из Excel таблицы, если он в ней есть
    row_num = row_map.get(i.chunk_id)
    cached_vec_text = ws.cell(row=row_num, column=vec_col_id).value if row_num else None

    # Если текст в таблице найден, используем его, извлекаем из него термины. Если текст не найден, создаем его.
    if cached_vec_text:
        context_text = cached_vec_text
        found_terms = find_terms_in_text(raw_plain_text, search_index, terminology)
    else:
        context_text, found_terms = create_chunk_text_context(
            chunk_tree, i.chunk_id, parent_chain,
            system_prompt, terminology, search_index,
            abbreviations, raw_plain_text
        )
        if row_num:
            ws.cell(row=row_num, column=vec_col_id, value=context_text)
            wb.save(file_path)

    result_chunk["chunk_text_context"] = context_text

    metadata["has_abbreviations"] = detect_abbreviations(plain_text, abbreviations)
    metadata["found_terms"] = found_terms
    metadata["parents_depth"] = parents_depth
    result_chunk["chunk_text_plain"] = plain_text
    result_chunk["metadata"] = metadata

    result_data.append(result_chunk)

print(f"\nDone! {total_chunks} chunks processed.")
new_logger.info(msg="Data object created")

try:
    save_to_json(result_data, output_path)
    print(f"Saved to {output_path}")
    new_logger.info(msg="JSON file saved")

except Exception as e:
    logging.error(msg=e)
    print(e)
    exit()

#TODO добавить режимы для добавления векторизованного текста: 1) не добавлять вообще
# 2) добавлять в поданную таблицу (нынешняя логика)
# 3) создавать новую таблицу