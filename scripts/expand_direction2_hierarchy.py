"""
expand_direction2_hierarchy.py — Раскрытие направление 2: статические кластеры банка.

МЕТОД:
  В отличие от направления 1 (кластеры формируются при поиске для каждого запроса),
  здесь банк кластеризуется ОДИН РАЗ по своим эмбеддингам. Это даёт стабильные
  смысловые группы, не зависящие от AS-IS запроса.

  Кластеризация банка на 98% совпадает с иерархической структурой таблицы,
  но может выявить неочевидные семантические связи между пунктами из РАЗНЫХ секций.

АЛГОРИТМ:
  1. Загружаем ВСЕ эмбеддинги банка из OpenSearch (processscout_bank)
  2. UMAP (50d, cosine) + HDBSCAN (min_cluster_size=3)
  3. Кластеры больше max_cluster_size (=7) разбиваются через KMeans
  4. Для каждого подтверждённого маппинга:
     - Находим кластер содержащий подтверждённый bank_id
     - Собираем всех соседей по этому статическому кластеру
  4. LLM (gpt-5.4-mini) оценивает соседей: нужен / контекст / не нужен
  5. Тексты из bank_4-5.xlsx

ОТЛИЧИЕ ОТ НАПРАВЛЕНИЯ 1:
  - Направление 1: кластеры зависят от AS-IS запроса (разные для narrative/clean/enriched)
  - Направление 2: кластеры фиксированные, зависят только от структуры банка
  - Направление 2 может найти связи которые направление 1 пропускает (и наоборот)

ЗАПУСК (на маке, обращается к OpenSearch на VM):
  python scripts/expand_direction2_hierarchy.py \\
    --bank bank_4-5.xlsx \\
    --output output/expansion_d2.json

  # Без LLM:
  python scripts/expand_direction2_hierarchy.py \\
    --bank bank_4-5.xlsx \\
    --output output/expansion_d2.json \\
    --no-llm

  # Другой min_cluster_size:
  python scripts/expand_direction2_hierarchy.py \\
    --bank bank_4-5.xlsx \\
    --min-cluster-size 10 \\
    --output output/expansion_d2.json
"""

import json
import argparse
import re
import numpy as np
from pathlib import Path
from collections import defaultdict


# ── Конфигурация ──
OPENSEARCH_URL      = "https://10.40.10.111:9200"
OPENSEARCH_USER     = "admin"
OPENSEARCH_PASSWORD = "ProcessScout_2026!"
INDEX_BANK          = "processscout_bank"

# HDBSCAN параметры
MIN_CLUSTER_SIZE    = 3      # минимальный размер кластера (для HDBSCAN)
MAX_CLUSTER_SIZE    = 7      # максимальный: кластеры больше этого будут разбиты
UMAP_N_COMPONENTS   = 50
UMAP_METRIC         = "cosine"

def load_confirmed_mappings(consolidated_path):
    """
    Reads consolidated_l3.json and extracts confirmed mappings.
    Returns dict: {pair_id: [(bank_id, bank_desc, operation), ...]}

    We take items with category "основной" as confirmed mappings.
    """
    with open(consolidated_path, encoding="utf-8") as f:
        data = json.load(f)

    mappings = {}
    for item in data.get("consolidated", []):
        pair_id = item["pair_id"]
        operation = item.get("operation", "")
        confirmed = []
        for section in item.get("sections", []):
            if section.get("category") == "основной":
                bank_id = section["representative_id"]
                # Extract short description from representative_text
                text = section.get("representative_text", "")
                # Take first line or first 100 chars
                desc = text.split("\n")[0][:100] if text else bank_id
                confirmed.append((bank_id, desc, operation))
        if confirmed:
            mappings[pair_id] = confirmed

    return mappings


# ═══════════════════════════════════════════════════════════════
# Утилиты
# ═══════════════════════════════════════════════════════════════

def extract_l3(bank_id: str) -> str:
    clean = bank_id.replace("FB-", "")
    clean = re.sub(r'[а-ед]\)$', '', clean)
    parts = clean.split(".")
    if len(parts) >= 3:
        return f"{parts[0]}.{parts[1]}.{parts[2]}"
    elif len(parts) >= 2:
        return f"{parts[0]}.{parts[1]}"
    return parts[0]


def load_bank_texts(xlsx_path: str) -> dict:
    """Загружает чистые формулировки из Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    texts = {}
    last_numbered_id = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        raw_id = str(row[0]).strip()
        text = str(row[1] or "").strip()
        if not text:
            continue
        is_letter = bool(re.match(r'^[а-ед]\)$', raw_id))
        if is_letter:
            if last_numbered_id:
                full_id = f"{last_numbered_id}{raw_id}"
            else:
                continue
        else:
            full_id = raw_id
            last_numbered_id = raw_id
        texts[f"FB-{full_id}"] = text
    return texts


# ═══════════════════════════════════════════════════════════════
# Загрузка эмбеддингов из OpenSearch
# ═══════════════════════════════════════════════════════════════

def load_bank_embeddings() -> tuple[list[str], np.ndarray]:
    """
    Загружает все эмбеддинги из processscout_bank.
    Возвращает (list_of_ids, numpy_array_of_embeddings).
    """
    from opensearchpy import OpenSearch

    client = OpenSearch(
        hosts=[OPENSEARCH_URL],
        http_auth=(OPENSEARCH_USER, OPENSEARCH_PASSWORD),
        use_ssl=True, verify_certs=False, ssl_show_warn=False, timeout=60
    )
    print(f"  OpenSearch: {client.info()['version']['number']}")

    ids = []
    embeddings = []

    resp = client.search(
        index=INDEX_BANK,
        body={"size": 500, "query": {"match_all": {}}},
        scroll="2m"
    )
    scroll_id = resp["_scroll_id"]
    hits = resp["hits"]["hits"]

    while hits:
        for h in hits:
            emb = h["_source"].get("embedding", [])
            if emb:
                ids.append(h["_id"])
                embeddings.append(emb)
        resp = client.scroll(scroll_id=scroll_id, scroll="2m")
        hits = resp["hits"]["hits"]

    try:
        client.clear_scroll(scroll_id=scroll_id)
    except Exception:
        pass

    emb_array = np.array(embeddings, dtype=np.float32)
    print(f"  Загружено эмбеддингов: {emb_array.shape}")
    return ids, emb_array


# ═══════════════════════════════════════════════════════════════
# Кластеризация банка
# ═══════════════════════════════════════════════════════════════

def cluster_bank(ids: list[str], embeddings: np.ndarray, min_cluster_size: int, max_cluster_size: int) -> dict:
    """
    UMAP + HDBSCAN кластеризация банка.
    Кластеры больше max_cluster_size разбиваются рекурсивно через KMeans.
    Возвращает dict {cluster_label: [list_of_bank_ids]}.
    """
    import umap
    import hdbscan
    from sklearn.cluster import KMeans

    print(f"  UMAP ({UMAP_N_COMPONENTS}d, {UMAP_METRIC})...")
    reducer = umap.UMAP(
        n_components=UMAP_N_COMPONENTS,
        metric=UMAP_METRIC,
        random_state=42,
        n_neighbors=15,
    )
    reduced = reducer.fit_transform(embeddings)
    print(f"    → {reduced.shape}")

    print(f"  HDBSCAN (min_cluster_size={min_cluster_size})...")
    clusterer = hdbscan.HDBSCAN(
        min_cluster_size=min_cluster_size,
        metric="euclidean",
        cluster_selection_method="eom",
    )
    labels = clusterer.fit_predict(reduced)

    # Группируем по кластерам (с индексами для доступа к эмбеддингам)
    raw_clusters = defaultdict(list)  # label → [(idx, bank_id)]
    for idx, label in enumerate(labels):
        raw_clusters[label].append((idx, ids[idx]))

    n_raw = len([k for k in raw_clusters if k >= 0])
    print(f"  До разбиения: {n_raw} кластеров")

    # Разбиваем кластеры > max_cluster_size
    final_clusters = {}
    next_label = 0
    split_count = 0

    for label, members in raw_clusters.items():
        if label == -1:
            # Outliers остаются как есть
            final_clusters[-1] = [bid for _, bid in members]
            continue

        if len(members) <= max_cluster_size:
            # Нормальный размер — оставляем
            final_clusters[next_label] = [bid for _, bid in members]
            next_label += 1
        else:
            # Слишком большой — разбиваем KMeans
            member_indices = [idx for idx, _ in members]
            member_ids = [bid for _, bid in members]
            member_embeddings = reduced[member_indices]

            # Количество подкластеров: ceil(размер / max_size)
            import math
            n_sub = math.ceil(len(members) / max_cluster_size)

            km = KMeans(n_clusters=n_sub, random_state=42, n_init=10)
            sub_labels = km.fit_predict(member_embeddings)

            for sub_label in range(n_sub):
                sub_members = [member_ids[i] for i in range(len(members))
                               if sub_labels[i] == sub_label]
                if sub_members:
                    final_clusters[next_label] = sub_members
                    next_label += 1
                    split_count += 1

    n_final = len([k for k in final_clusters if k >= 0])
    n_outliers = len(final_clusters.get(-1, []))
    sizes = [len(v) for k, v in final_clusters.items() if k >= 0]

    print(f"  После разбиения (max={max_cluster_size}): {n_final} кластеров "
          f"({split_count} подкластеров создано), {n_outliers} outliers")
    if sizes:
        print(f"    Размеры: min={min(sizes)}, max={max(sizes)}, "
              f"median={sorted(sizes)[len(sizes)//2]}, mean={sum(sizes)/len(sizes):.1f}")

    return final_clusters


# ═══════════════════════════════════════════════════════════════
# LLM оценка
# ═══════════════════════════════════════════════════════════════

def llm_evaluate(bp_operation, confirmed_id, confirmed_desc, neighbors):
    """LLM оценивает соседей по статическому кластеру."""
    from openai import OpenAI
    client = OpenAI()

    docs = ""
    for i, n in enumerate(neighbors[:15], 1):
        docs += (
            f"\n--- Кандидат {i} (ID: {n['id']}, секция L3={n['l3']}) ---\n"
            f"Текст: {n['text'][:200]}\n"
        )

    prompt = f"""Ты — эксперт по автоматизации производственных предприятий (PLM/ERP/APS/MES).

КОНТЕКСТ: Аналитик подтвердил что шаг бизнес-процесса соответствует конкретному пункту банка функциональности. Теперь мы ищем какие ЕЩЁ функции из банка могут понадобиться заказчику в связи с этим маппингом.

ШАГ БИЗНЕС-ПРОЦЕССА:
{bp_operation}

ПОДТВЕРЖДЁННЫЙ МАППИНГ:
{confirmed_id}: {confirmed_desc}

КАНДИДАТЫ НА РАСКРЫТИЕ (из того же семантического кластера банка):
{docs}

Для каждого кандидата определи:
- "нужен" — функция НАПРЯМУЮ дополняет подтверждённый маппинг. Объясни: что именно она добавляет и зачем заказчику.
- "контекст" — тематически связана, но прямой необходимости нет.
- "не нужен" — не добавляет ценности.

ПРАВИЛА ОЦЕНКИ:
1. Если формулировка подтверждённого маппинга ССЫЛАЕТСЯ на понятие, которое описывается в кандидате — это "нужен". Пример: маппинг "КСГ производственных и закупочных заказов" содержит "производственные заказы" → кандидат "Формирование производственных заказов" нужен, потому что раскрывает что такое "производственные заказы" в контексте маппинга.
2. Если кандидат раскрывает КОНКРЕТНЫЙ АСПЕКТ подтверждённого маппинга (подэтап, параметр, настройку) — это "нужен".
3. Родительские пункты (более общие чем подтверждённый маппинг) — обычно "не нужен".
4. Пункты из совсем другой функциональной области — "не нужен".
5. Лучше пропустить полезный пункт чем включить мусор. Но не перестраховывайся: если видишь прямую связь по формулировке — ставь "нужен".

Объяснение для "нужен" ОБЯЗАТЕЛЬНО: ЧТО раскрывает и ЗАЧЕМ заказчику.
Объяснение для "не нужен" не нужно.

JSON: {{"results": [{{"id": "...", "verdict": "нужен|контекст|не нужен", "explanation": "..."}}]}}"""

    try:
        resp = client.responses.create(model="gpt-5.4-mini", input=prompt)
        raw = resp.output_text.strip()
        raw = re.sub(r'```json\s*', '', raw)
        raw = re.sub(r'```\s*', '', raw)
        parsed = json.loads(raw)
        return {r["id"]: r for r in parsed.get("results", [])}
    except Exception as e:
        print(f"    LLM ошибка: {e}")
        return {}


# ═══════════════════════════════════════════════════════════════
# Основной пайплайн
# ═══════════════════════════════════════════════════════════════

def run(bank_path, min_cluster_size, max_cluster_size, use_llm, confirmed_mappings):
    # 1. Загружаем тексты банка
    print("Загрузка текстов банка...")
    bank_texts = load_bank_texts(bank_path)
    print(f"  Текстов: {len(bank_texts)}")

    # 2. Загружаем эмбеддинги
    print("Загрузка эмбеддингов из OpenSearch...")
    ids, embeddings = load_bank_embeddings()

    # 3. Кластеризация
    print("Кластеризация банка...")
    clusters = cluster_bank(ids, embeddings, min_cluster_size, max_cluster_size)

    # 4. Строим обратный индекс: bank_id → cluster_label
    id_to_cluster = {}
    for label, members in clusters.items():
        for mid in members:
            id_to_cluster[mid] = label

    # 5. Показываем кластеры для подтверждённых пунктов
    print(f"\nКластеры подтверждённых пунктов:")
    for pair_id, mappings in confirmed_mappings.items():
        for bank_id, desc, _op in mappings:
            cl = id_to_cluster.get(bank_id, "???")
            cl_size = len(clusters.get(cl, []))
            print(f"  {bank_id}: cluster_{cl} ({cl_size} пунктов)")

    # 6. Раскрытие
    print(f"\n{'='*60}")
    print("РАСКРЫТИЕ:")

    all_expansions = []

    for pair_id, mappings in confirmed_mappings.items():
        for bank_id, bank_desc, bp_op in mappings:
            cl_label = id_to_cluster.get(bank_id)
            if cl_label is None or cl_label == -1:
                print(f"\n{pair_id} → {bank_id}: OUTLIER (не в кластере)")
                all_expansions.append({
                    "pair_id": pair_id,
                    "confirmed_id": bank_id,
                    "confirmed_desc": bank_desc,
                    "operation": bp_op,
                    "cluster_label": int(cl_label) if cl_label is not None else -1,
                    "cluster_size": 0,
                    "neighbors_found": 0,
                    "neighbors": [],
                })
                continue

            cl_members = clusters[cl_label]
            neighbors_ids = [m for m in cl_members if m != bank_id]

            print(f"\n{pair_id} → {bank_id}: cluster_{cl_label} ({len(cl_members)} пунктов, {len(neighbors_ids)} соседей)")

            # Формируем список соседей с текстами
            neighbors = []
            for nid in neighbors_ids:
                neighbors.append({
                    "id": nid,
                    "l3": extract_l3(nid),
                    "text": bank_texts.get(nid, nid),
                })

            # Показываем
            for n in neighbors[:10]:
                print(f"    {n['id']}: {n['text'][:70]}")

            # LLM оценка
            if use_llm and neighbors:
                print(f"  LLM оценка...")
                verdicts = llm_evaluate(bp_op, bank_id, bank_desc, neighbors)
                for n in neighbors:
                    v = verdicts.get(n["id"], {})
                    n["llm_verdict"] = v.get("verdict", "")
                    n["llm_explanation"] = v.get("explanation", "")

                n_needed = sum(1 for n in neighbors if n.get("llm_verdict") == "нужен")
                n_ctx = sum(1 for n in neighbors if n.get("llm_verdict") == "контекст")
                print(f"    → нужен: {n_needed}, контекст: {n_ctx}")

            all_expansions.append({
                "pair_id": pair_id,
                "confirmed_id": bank_id,
                "confirmed_desc": bank_desc,
                "operation": bp_op,
                "cluster_label": int(cl_label),
                "cluster_size": len(cl_members),
                "neighbors_found": len(neighbors),
                "neighbors": neighbors,
            })

    return all_expansions


def main():
    parser = argparse.ArgumentParser(
        description="Раскрытие направление 2: статические кластеры банка"
    )
    parser.add_argument("--bank", default="bank_4-5.xlsx",
                        help="Путь к bank_45.xlsx")
    parser.add_argument("--output", default="output/expansion_d2.json")
    parser.add_argument("--min-cluster-size", type=int, default=MIN_CLUSTER_SIZE,
                        help=f"Мин. размер кластера HDBSCAN (по умолч.: {MIN_CLUSTER_SIZE})")
    parser.add_argument("--max-cluster-size", type=int, default=MAX_CLUSTER_SIZE,
                        help=f"Макс. размер: кластеры больше будут разбиты (по умолч.: {MAX_CLUSTER_SIZE})")
    parser.add_argument("--consolidated", required=True,
                        help="Путь к consolidated_l3.json с подтверждёнными маппингами")
    parser.add_argument("--no-llm", action="store_true")
    args = parser.parse_args()

    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    confirmed_mappings = load_confirmed_mappings(args.consolidated)
    print(f"Загружено маппингов: {sum(len(v) for v in confirmed_mappings.values())} "
          f"из {len(confirmed_mappings)} шагов БП")

    expansions = run(args.bank, args.min_cluster_size, args.max_cluster_size,
                     use_llm=not args.no_llm, confirmed_mappings=confirmed_mappings)

    # Сохранение
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({
            "metadata": {
                "direction": 2,
                "method": "static_hdbscan",
                "min_cluster_size": args.min_cluster_size,
                "max_cluster_size": args.max_cluster_size,
                "index": INDEX_BANK,
            },
            "expansions": expansions,
        }, f, indent=2, ensure_ascii=False)
    print(f"\nСохранено: {out_path}")

    # Сводка
    print(f"\n{'='*60}")
    print(f"СВОДКА направление 2 (статические кластеры, min={args.min_cluster_size}, max={args.max_cluster_size}):")
    for exp in expansions:
        n_total = exp["neighbors_found"]
        n_needed = sum(1 for n in exp["neighbors"] if n.get("llm_verdict") == "нужен")
        n_ctx = sum(1 for n in exp["neighbors"] if n.get("llm_verdict") == "контекст")
        print(f"  {exp['pair_id']} → {exp['confirmed_id']}: "
              f"cluster_{exp['cluster_label']} ({exp['cluster_size']}), "
              f"{n_total} соседей, {n_needed} нужен, {n_ctx} контекст")


if __name__ == "__main__":
    main()