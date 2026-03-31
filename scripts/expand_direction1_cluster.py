"""
expand_direction1_cluster.py — Раскрытие направление 1: динамические кластеры.

МЕТОД:
  При маппинге каждый AS-IS чанк проходит через UMAP+HDBSCAN кластеризацию,
  формируя семантические кластеры из пунктов банка. Подтверждённый аналитиком
  пункт находится в одном из этих кластеров — рядом с семантически близкими соседями.

  Направление 1 собирает ВСЕХ соседей подтверждённого пункта из ВСЕХ кластеров
  (по всем типам чанков: narrative, enriched, clean). Чем в большем количестве
  кластеров сосед встречается рядом с подтверждённым пунктом — тем сильнее сигнал.

АЛГОРИТМ:
  1. Для каждого подтверждённого маппинга (pair_id → bank_id):
     - Проходим по всем chunk types (narrative, enriched, clean)
     - В каждом находим кластер содержащий подтверждённый bank_id
     - Собираем всех соседей из этого кластера
  2. Дедупликация: считаем в скольких кластерах каждый сосед встречается
  3. Ранжирование: по количеству кластеров (убывание), затем по количеству типов чанков
  4. LLM (gpt-5.4-mini) оценивает top-15 соседей: нужен / контекст / не нужен
  5. Тексты соседей загружаются из bank_4-5.xlsx (не из результатов поиска)

ОГРАНИЧЕНИЯ:
  - Cosine similarity соседей к запросу НЕ используется. Соседи попали в кластер
    через UMAP+HDBSCAN, а не через прямой cosine. Ранжирование — по частоте
    совместного появления в кластерах.
  - LLM оценивает только top-15 соседей. При большом количестве соседей (>30)
    часть остаётся без оценки.

ИДЕЯ НА БУДУЩЕЕ:
  Добавить cosine similarity соседей к AS-IS запросу как дополнительный сигнал.
  Для этого нужно: (1) сохранять эмбеддинги соседей при кластеризации, или
  (2) делать дополнительный KNN запрос для каждого соседа. Это улучшит
  ранжирование, но увеличит время работы.

ЗАПУСК:
  python scripts/expand_direction1_cluster.py \\
    --results output/best_result_v1/results_clustered.json \\
    --bank bank_4-5.xlsx \\
    --output output/expansion_d1.json

  # Без LLM (только сбор соседей):
  python scripts/expand_direction1_cluster.py \\
    --results output/best_result_v1/results_clustered.json \\
    --bank bank_4-5.xlsx \\
    --output output/expansion_d1.json \\
    --no-llm
"""

import json
import argparse
import re
from pathlib import Path
from collections import defaultdict


def load_confirmed_mappings(consolidated_path):
    """
    Reads consolidated_l3.json and extracts confirmed mappings.
    Returns dict: {pair_id: [(bank_id, bank_desc, operation), ...]}

    We take items with category "основной" as confirmed mappings.
    """
    with open(consolidated_path, encoding="utf-8") as f:
        data = json.load(f)

    mappings = {}
    for item in data.get("consolidated", []):
        pair_id = item["pair_id"]
        operation = item.get("operation", "")
        confirmed = []
        for section in item.get("sections", []):
            if section.get("category") == "основной":
                bank_id = section["representative_id"]
                text = section.get("representative_text", "")
                desc = text.split("\n")[0][:100] if text else bank_id
                confirmed.append((bank_id, desc, operation))
        if confirmed:
            mappings[pair_id] = confirmed

    return mappings


def extract_l3(bank_id: str) -> str:
    """FB-4.7.3 → 4.7.3"""
    clean = bank_id.replace("FB-", "")
    clean = re.sub(r'[а-ед]\)$', '', clean)
    parts = clean.split(".")
    if len(parts) >= 3:
        return f"{parts[0]}.{parts[1]}.{parts[2]}"
    elif len(parts) >= 2:
        return f"{parts[0]}.{parts[1]}"
    return parts[0]


def load_bank_texts(xlsx_path: str) -> dict:
    """
    Загружает чистые формулировки банковских пунктов из Excel.
    Колонка 1 = ID, колонка 2 = текст.
    Буквенные пункты (а, б) → составной ID: "FB-4.6.2.6а)"
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    texts = {}
    last_numbered_id = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        raw_id = str(row[0]).strip()
        text = str(row[1] or "").strip()
        if not text:
            continue
        is_letter = bool(re.match(r'^[а-ед]\)$', raw_id))
        if is_letter:
            if last_numbered_id:
                full_id = f"{last_numbered_id}{raw_id}"
            else:
                continue
        else:
            full_id = raw_id
            last_numbered_id = raw_id
        texts[f"FB-{full_id}"] = text
    return texts


def find_cluster_neighbors(results: list[dict], pair_id: str, bank_id: str) -> dict:
    """
    Для заданного pair_id и bank_id находит всех соседей по кластерам.
    
    Проходит по всем chunk types (narrative, enriched, clean).
    Для каждого: находит кластер содержащий bank_id, собирает соседей.
    
    Возвращает dict {neighbor_id: {count, chunk_types, cluster_names, cluster_sizes}}.
    """
    neighbors = defaultdict(lambda: {
        "count": 0,
        "chunk_types": [],
        "cluster_names": [],
        "cluster_sizes": [],
    })

    for r in results:
        # Фильтр по pair_id
        if r.get("chunk_meta", {}).get("pair_id") != pair_id:
            continue
        if r.get("status") != "ok":
            continue

        chunk_id = r["chunk_id"]
        chunk_type = r.get("chunk_meta", {}).get("chunk_type", "?")
        clusters = r.get("metadata", {}).get("clusters_formed", {})

        # Ищем кластер содержащий bank_id
        for cl_name, items in clusters.items():
            if bank_id in items:
                # Нашли! Собираем всех соседей
                for item_id in items:
                    if item_id == bank_id:
                        continue  # пропускаем сам подтверждённый пункт
                    n = neighbors[item_id]
                    n["count"] += 1
                    n["chunk_types"].append(chunk_type)
                    n["cluster_names"].append(f"{chunk_id}:{cl_name}")
                    n["cluster_sizes"].append(len(items))
                break  # bank_id может быть только в одном кластере на чанк

    return dict(neighbors)


def get_bank_text(results: list[dict], bank_id: str) -> str:
    """
    Извлекает текст банковского пункта из результатов поиска.
    Ищет chunk_text_plain в любом хите содержащем bank_id.
    """
    for r in results:
        if r.get("status") != "ok":
            continue
        meta = r.get("metadata", {})
        clusters = meta.get("clusters_formed", {})
        for cl_name, items in clusters.items():
            if bank_id in items:
                # Нашли пункт — теперь нужен его текст
                # Текст может быть в step1_details или в кластерных документах
                # Проще: ищем в step1_details
                for d in meta.get("step1_details", []):
                    if d.get("best_id") == bank_id:
                        return d.get("explanation", "")[:200]
                return ""
    return ""


def llm_evaluate_neighbors(
    bp_operation: str,
    confirmed_id: str,
    confirmed_desc: str,
    neighbors: list[dict],
    openai_key: str = None,
) -> list[dict]:
    """
    LLM оценивает соседей: нужны ли они заказчику как дополнение.
    """
    from openai import OpenAI
    client = OpenAI()

    docs = ""
    for i, n in enumerate(neighbors[:15], 1):
        docs += (
            f"\n--- Сосед {i} (ID: {n['id']}, секция L3={n['l3']}, "
            f"встречается в {n['count']} кластерах) ---\n"
            f"Текст: {n['text'][:200]}\n"
        )

    prompt = f"""Ты — эксперт по автоматизации производственных предприятий (PLM/ERP/APS/MES).

КОНТЕКСТ: Аналитик подтвердил что шаг бизнес-процесса соответствует конкретному пункту банка функциональности. Теперь мы ищем какие ЕЩЁ функции из банка могут понадобиться заказчику в связи с этим маппингом.

ШАГ БИЗНЕС-ПРОЦЕССА:
{bp_operation}

ПОДТВЕРЖДЁННЫЙ МАППИНГ:
{confirmed_id}: {confirmed_desc}

КАНДИДАТЫ НА РАСКРЫТИЕ:
{docs}

Для каждого кандидата определи:
- "нужен" — функция НАПРЯМУЮ дополняет подтверждённый маппинг. Объясни: что именно она добавляет и зачем заказчику.
- "контекст" — тематически связана, но прямой необходимости нет.
- "не нужен" — не добавляет ценности.

ПРАВИЛА ОЦЕНКИ:
1. Если формулировка подтверждённого маппинга ССЫЛАЕТСЯ на понятие, которое описывается в кандидате — это "нужен". Пример: маппинг "КСГ производственных и закупочных заказов" содержит "производственные заказы" → кандидат "Формирование производственных заказов" нужен, потому что раскрывает что такое "производственные заказы" в контексте маппинга.
2. Если кандидат раскрывает КОНКРЕТНЫЙ АСПЕКТ подтверждённого маппинга (подэтап, параметр, настройку) — это "нужен".
3. Родительские пункты (более общие чем подтверждённый маппинг) — обычно "не нужен".
4. Пункты из совсем другой функциональной области — "не нужен".
5. Лучше пропустить полезный пункт чем включить мусор. Но не перестраховывайся: если видишь прямую связь по формулировке — ставь "нужен".

Объяснение для "нужен" ОБЯЗАТЕЛЬНО: ЧТО раскрывает и ЗАЧЕМ заказчику.
Объяснение для "не нужен" не нужно.

JSON: {{"results": [{{"id": "...", "verdict": "нужен|контекст|не нужен", "explanation": "..."}}]}}"""

    try:
        resp = client.responses.create(model="gpt-5.4-mini", input=prompt)
        raw = resp.output_text.strip()
        raw = re.sub(r'```json\s*', '', raw)
        raw = re.sub(r'```\s*', '', raw)
        parsed = json.loads(raw)
        return parsed.get("results", [])
    except Exception as e:
        print(f"    LLM ошибка: {e}")
        return []


def run_expansion(results_path: str, use_llm: bool = True, bank_path: str = "bank_4-5.xlsx",
                   confirmed_mappings: dict = None) -> list[dict]:
    """Основной пайплайн раскрытия по направлению 1."""

    with open(results_path, encoding="utf-8") as f:
        data = json.load(f)
    results = data["results"]
    print(f"Загружено результатов: {len(results)}")

    # Загружаем тексты банка
    bank_texts = load_bank_texts(bank_path)
    print(f"Загружено текстов банка: {len(bank_texts)}")

    all_expansions = []

    for pair_id, mappings in confirmed_mappings.items():
        for bank_id, bank_desc, bp_operation in mappings:
            print(f"\n{'='*60}")
            print(f"{pair_id} → {bank_id}: {bank_desc}")
            print(f"AS-IS: {bp_operation[:80]}")

            # 1. Собираем соседей из кластеров
            neighbors_raw = find_cluster_neighbors(results, pair_id, bank_id)
            print(f"  Соседей найдено: {len(neighbors_raw)}")

            if not neighbors_raw:
                print(f"  Нет соседей!")
                all_expansions.append({
                    "pair_id": pair_id,
                    "confirmed_id": bank_id,
                    "confirmed_desc": bank_desc,
                    "operation": bp_operation,
                    "neighbors_found": 0,
                    "neighbors": [],
                })
                continue

            # 2. Ранжируем: чем больше кластеров содержат соседа — тем сильнее сигнал
            neighbors_list = []
            for nid, ndata in neighbors_raw.items():
                neighbors_list.append({
                    "id": nid,
                    "count": ndata["count"],
                    "chunk_types": list(set(ndata["chunk_types"])),
                    "n_chunk_types": len(set(ndata["chunk_types"])),
                    "cluster_names": ndata["cluster_names"],
                    "avg_cluster_size": round(sum(ndata["cluster_sizes"]) / len(ndata["cluster_sizes"]), 1),
                    "l3": extract_l3(nid),
                    "text": "",  # заполним позже из банка если нужно
                })

            # Сортируем: по количеству кластеров (убывание), потом по количеству типов чанков
            neighbors_list.sort(key=lambda x: (-x["count"], -x["n_chunk_types"]))

            # Показываем top-10
            for i, n in enumerate(neighbors_list[:10]):
                print(f"    #{i+1}: {n['id']} (в {n['count']} кластерах, "
                      f"{n['n_chunk_types']} типов, L3={n['l3']})")

            # 3. LLM оценка
            if use_llm and neighbors_list:
                print(f"  LLM оценка top-15...")

                # Загружаем тексты соседей из банка
                for n in neighbors_list:
                    n["text"] = bank_texts.get(n["id"], n["id"])

                verdicts = llm_evaluate_neighbors(
                    bp_operation, bank_id, bank_desc,
                    neighbors_list, 
                )

                verdicts_map = {v["id"]: v for v in verdicts}
                for n in neighbors_list:
                    v = verdicts_map.get(n["id"], {})
                    n["llm_verdict"] = v.get("verdict", "")
                    n["llm_explanation"] = v.get("explanation", "")
                    if n["llm_verdict"]:
                        if n["id"] in [x["id"] for x in neighbors_list[:10]]:
                            print(f"    {n['id']}: {n['llm_verdict']}")

            all_expansions.append({
                "pair_id": pair_id,
                "confirmed_id": bank_id,
                "confirmed_desc": bank_desc,
                "operation": bp_operation,
                "neighbors_found": len(neighbors_list),
                "neighbors": neighbors_list,
            })

    return all_expansions


def main():
    parser = argparse.ArgumentParser(
        description="Раскрытие направление 1: соседи из динамических кластеров"
    )
    parser.add_argument("--results", required=True,
                        help="Путь к results_clustered.json")
    parser.add_argument("--output", default="output/expansion_d1.json")
    parser.add_argument("--no-llm", action="store_true",
                        help="Без LLM оценки (только сбор соседей)")
    parser.add_argument("--bank", default="bank_4-5.xlsx",
                        help="Путь к bank_45.xlsx для загрузки текстов")
    parser.add_argument("--consolidated", required=True,
                        help="Путь к consolidated_l3.json с подтверждёнными маппингами")
    args = parser.parse_args()

    # Загружаем .env для OpenAI
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    confirmed_mappings = load_confirmed_mappings(args.consolidated)
    print(f"Загружено маппингов: {sum(len(v) for v in confirmed_mappings.values())} "
          f"из {len(confirmed_mappings)} шагов БП")

    expansions = run_expansion(args.results, use_llm=not args.no_llm, bank_path=args.bank,
                               confirmed_mappings=confirmed_mappings)

    # Сохранение
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(expansions, f, indent=2, ensure_ascii=False)
    print(f"\nСохранено: {out_path}")

    # Сводка
    print(f"\n{'='*60}")
    print("СВОДКА направление 1 (динамические кластеры):")
    for exp in expansions:
        n_total = exp["neighbors_found"]
        n_needed = sum(1 for n in exp["neighbors"] if n.get("llm_verdict") == "нужен")
        n_context = sum(1 for n in exp["neighbors"] if n.get("llm_verdict") == "контекст")
        print(f"  {exp['pair_id']} → {exp['confirmed_id']}: "
              f"{n_total} соседей, {n_needed} нужен, {n_context} контекст")


if __name__ == "__main__":
    main()