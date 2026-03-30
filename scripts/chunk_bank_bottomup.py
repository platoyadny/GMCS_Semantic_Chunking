"""
chunk_bank_bottomup.py — Bottom-up чанкинг банка функциональности (v3).

Запуск:
  python scripts/chunk_bank_bottomup.py --input bank_4-5.xlsx
  python scripts/chunk_bank_bottomup.py --input bank_4-5.xlsx --no-llm
  python scripts/chunk_bank_bottomup.py --input bank_4-5.xlsx --resume
"""

import json, argparse, re, sys
from pathlib import Path
from collections import OrderedDict, defaultdict

MAX_DEPTH = 4
LLM_CACHE_FILE = "output/cache/llm_reformulations_cache.json"
OUTPUT_FILE = "output/bank_variants.json"

def parse_hierarchy(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    items = OrderedDict()
    last_numbered_id = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        raw_id = row[0]
        if raw_id is None: continue
        raw_id = str(raw_id).strip()
        clean_text = str(row[1] or "").strip()
        if not clean_text: continue
        is_letter = bool(re.match(r'^[а-ед]\)$', raw_id))
        if is_letter:
            if last_numbered_id is None: continue
            full_id = f"{last_numbered_id}{raw_id}"
            parent_id = last_numbered_id
        else:
            full_id = raw_id
            last_numbered_id = raw_id
            parts = full_id.split(".")
            if len(parts) >= 2: parent_id = ".".join(parts[:-1])
            elif full_id in ("4","5"): parent_id = None
            else: parent_id = None
        if full_id in ("4","5"): level = 1
        elif is_letter: level = len(last_numbered_id.split(".")) + 2
        else: level = len(full_id.split(".")) + 1
        items[full_id] = {"id": full_id, "clean_text": clean_text, "parent_id": parent_id, "level": level, "is_letter": is_letter, "children_ids": []}
    for item_id, item in items.items():
        chain = []
        pid = item["parent_id"]
        while pid and pid in items:
            chain.insert(0, pid)
            pid = items[pid]["parent_id"]
        item["parent_chain"] = chain
        if item["parent_id"] and item["parent_id"] in items:
            items[item["parent_id"]]["children_ids"].append(item_id)
    for item in items.values():
        item["is_leaf"] = len(item["children_ids"]) == 0
    by_level = defaultdict(int)
    letters = 0
    for item in items.values():
        by_level[item["level"]] += 1
        if item["is_letter"]: letters += 1
    print(f"Распарсено пунктов: {len(items)} ({letters} буквенных)")
    for lvl in sorted(by_level.keys()):
        leaf_count = sum(1 for i in items.values() if i["level"] == lvl and i["is_leaf"])
        print(f"  Уровень {lvl}: {by_level[lvl]} ({leaf_count} листьев)")
    return items

def load_cache():
    p = Path(LLM_CACHE_FILE)
    if p.exists():
        with open(p, encoding="utf-8") as f: c = json.load(f)
        print(f"  LLM кэш: {len(c)} записей")
        return c
    return {}

def save_cache(cache):
    Path(LLM_CACHE_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(LLM_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)

def llm_call(prompt):
    from openai import OpenAI
    client = OpenAI()
    try:
        resp = client.responses.create(model="gpt-5.4-mini", input=prompt)
        return resp.output_text.strip().strip('"').strip("'")
    except Exception as e:
        print(f"      LLM ошибка: {e}")
        return ""

def make_letter_d0_prompt(item_text, parent_text):
    return f'''У тебя есть короткое название подпункта из справочника функций производственной системы и его родительский пункт. Короткое название без контекста бессмысленно.

ПОДПУНКТ: "{item_text}"
РОДИТЕЛЬСКИЙ ПУНКТ: "{parent_text}"

Сформулируй ОДНО предложение (10-20 слов), которое:
- Превращает короткое название в полноценное описание функции
- Использует контекст родительского пункта для придания смысла
- Звучит как описание в техническом справочнике, а не как заголовок

Примеры:
  Подпункт: "Частота поставок"
  Родитель: "Учёт планово-справочных параметров клиентов объектов спроса"
  → "Учёт частоты поставок при планировании под параметры клиентов"

  Подпункт: "На складе (свободные)"
  Родитель: "Детализация запасов по статусу хранения на складах"
  → "Учёт свободных запасов на складе при детализации по статусу хранения"

  Подпункт: "В пути"
  Родитель: "Детализация запасов по статусу хранения на складах"
  → "Учёт запасов в пути при детализации по статусу хранения на складах"

  Подпункт: "Группировка по толщине материала"
  Родитель: "Настройка правил формирования плановых партий по технологическим критериям"
  → "Группировка плановых партий по толщине материала как технологический критерий формирования"

  Подпункт: "Жёсткое распределение (резерв запасов только под конкретный заказ)"
  Родитель: "Поддержка типов распределения запасов"
  → "Жёсткое распределение с резервированием запасов только под конкретный заказ"

Ответь ТОЛЬКО результатом, без пояснений.'''

def make_enrich_prompt(prev_result, parent_text):
    return f'''У тебя есть формулировка функции производственной системы и контекст вышестоящего раздела. Нужно дополнить формулировку так, чтобы было понятно К КАКОЙ ОБЛАСТИ она относится.

ФОРМУЛИРОВКА:
"{prev_result}"

ВЫШЕСТОЯЩИЙ РАЗДЕЛ:
"{parent_text}"

Добавь к формулировке короткий смысловой хвост (2-7 слов), который отвечает на вопрос «в рамках чего?» или «для чего?», используя контекст вышестоящего раздела.

Правила:
- НЕ переписывай формулировку — только добавь хвост
- НЕ повторяй слова, уже имеющиеся в формулировке
- Хвост должен быть естественным продолжением предложения
- Результат — ОДНО предложение

Примеры:
  Формулировка: "Производственные календари и графики сменности"
  Раздел: "Учёт фонда рабочего времени при расчёте загрузки"
  → "Производственные календари и графики сменности при расчёте загрузки мощностей"

  Формулировка: "Производственные календари и графики сменности при расчёте загрузки мощностей"
  Раздел: "Планирование с учётом ограниченной мощности узких мест"
  → "Производственные календари и графики сменности при расчёте загрузки мощностей для планирования"

  Формулировка: "Формирование календарно-сетевого графика взаимосвязанных заказов"
  Раздел: "Формирование директивного плана"
  → "Формирование календарно-сетевого графика взаимосвязанных заказов в рамках директивного плана"

  Формулировка: "Учёт свободных запасов на складе при детализации по статусу хранения"
  Раздел: "Учёт текущего состояния производства (НЗП) и доступных ресурсов"
  → "Учёт свободных запасов на складе при детализации по статусу хранения в рамках НЗП"

  Формулировка: "Поддержка расчёта потребности в персонале по категориям и сменам"
  Раздел: "Планирование по организационным уровням"
  → "Поддержка расчёта потребности в персонале по категориям и сменам на уровне организации"

Ответь ТОЛЬКО результатом, без пояснений.'''

def template_letter_d0(item_text, parent_text):
    return f"{item_text} — {parent_text}"

def template_enrich(prev_result, parent_text):
    prev_words = set(prev_result.lower().split())
    parent_words = parent_text.split()
    new_words = [w for w in parent_words if w.lower() not in prev_words][:5]
    tail = " ".join(new_words) if new_words else parent_text[:30]
    return f"{prev_result} ({tail})"

def run_bottomup(items, use_llm, resume):
    cache = load_cache() if resume else {}
    all_levels = sorted(set(item["level"] for item in items.values()), reverse=True)
    print(f"\nУровни (от глубокого): {all_levels}")
    results = []
    total_llm = 0
    total_cached = 0
    for level in all_levels:
        level_items = [item for item in items.values() if item["level"] == level]
        print(f"\n{'='*60}")
        print(f"УРОВЕНЬ {level}: {len(level_items)} пунктов")
        print(f"{'='*60}")
        for idx, item in enumerate(level_items):
            bid = item["id"]
            clean = item["clean_text"]
            is_letter = item["is_letter"]
            variants = []
            chain_bottom_up = list(reversed(item["parent_chain"]))
            if is_letter and chain_bottom_up:
                nearest_parent_id = chain_bottom_up[0]
                nearest_parent = items.get(nearest_parent_id)
                if nearest_parent:
                    ck = f"{bid}__d0_letter"
                    if ck in cache:
                        d0_text = cache[ck]; total_cached += 1
                    elif use_llm:
                        d0_text = llm_call(make_letter_d0_prompt(clean, nearest_parent["clean_text"]))
                        if not d0_text: d0_text = template_letter_d0(clean, nearest_parent["clean_text"])
                        cache[ck] = d0_text; total_llm += 1
                    else:
                        d0_text = template_letter_d0(clean, nearest_parent["clean_text"])
                    variants.append({"depth": 0, "text": d0_text})
                    prev_result = d0_text
                    remaining = chain_bottom_up[1:]
                    for di, pid in enumerate(remaining[:MAX_DEPTH], 1):
                        pi = items.get(pid)
                        if not pi: break
                        ck = f"{bid}__d{di}"
                        if ck in cache:
                            ref = cache[ck]; total_cached += 1
                        elif use_llm:
                            ref = llm_call(make_enrich_prompt(prev_result, pi["clean_text"]))
                            if not ref: ref = template_enrich(prev_result, pi["clean_text"])
                            cache[ck] = ref; total_llm += 1
                        else:
                            ref = template_enrich(prev_result, pi["clean_text"])
                        variants.append({"depth": di, "text": ref, "parent_id": pid})
                        prev_result = ref
                else:
                    variants.append({"depth": 0, "text": clean})
            else:
                variants.append({"depth": 0, "text": clean})
                prev_result = clean
                for di, pid in enumerate(chain_bottom_up[:MAX_DEPTH], 1):
                    pi = items.get(pid)
                    if not pi: break
                    ck = f"{bid}__d{di}"
                    if ck in cache:
                        ref = cache[ck]; total_cached += 1
                    elif use_llm:
                        ref = llm_call(make_enrich_prompt(prev_result, pi["clean_text"]))
                        if not ref: ref = template_enrich(prev_result, pi["clean_text"])
                        cache[ck] = ref; total_llm += 1
                    else:
                        ref = template_enrich(prev_result, pi["clean_text"])
                    variants.append({"depth": di, "text": ref, "parent_id": pid})
                    prev_result = ref
            results.append({"bank_id": bid, "level": item["level"], "is_leaf": item["is_leaf"], "is_letter": is_letter, "clean_text": clean, "parent_id": item["parent_id"], "parent_chain": item["parent_chain"], "children_ids": item["children_ids"], "variants": variants})
            if use_llm and total_llm > 0 and total_llm % 50 == 0:
                save_cache(cache)
                print(f"    [{total_llm} LLM, {total_cached} кэш]")
            if idx < 3 or (idx+1) % 30 == 0:
                tag = " [БУКВА]" if is_letter else ""
                print(f"  [{idx+1}/{len(level_items)}] {bid}{tag}: {clean[:50]}")
                for v in variants:
                    print(f"    d{v['depth']}: {v['text'][:80]}")
    if use_llm and total_llm > 0: save_cache(cache)
    tv = sum(len(r["variants"]) for r in results)
    print(f"\n{'='*60}")
    print(f"Итого: {len(results)} пунктов, {tv} вариантов")
    print(f"  LLM вызовов: {total_llm}, из кэша: {total_cached}")
    return results

def main():
    parser = argparse.ArgumentParser(description="Bottom-up чанкинг банка (v3)")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", default=OUTPUT_FILE)
    parser.add_argument("--no-llm", action="store_true")
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()
    print(f"Парсинг {args.input}...")
    items = parse_hierarchy(args.input)
    results = run_bottomup(items, use_llm=not args.no_llm, resume=args.resume)
    print(f"\nПримеры:")
    for bid in ["4.6.2.6а)", "4.1.1а)", "4.6.8.3.1а)", "4.10.1.1", "5.13.1", "4.7.3", "4.7", "4.3"]:
        r = next((r for r in results if r["bank_id"] == bid), None)
        if r:
            tag = " [БУКВА]" if r["is_letter"] else ""
            print(f"\n  {bid}{tag} (level={r['level']}):")
            for v in r["variants"]:
                print(f"    d{v['depth']}: {v['text'][:90]}")
    output = {"metadata": {"total_items": len(results), "total_variants": sum(len(r["variants"]) for r in results), "max_depth": MAX_DEPTH, "levels": sorted(set(r["level"] for r in results), reverse=True)}, "items": results}
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\nСохранено: {out_path}")
    print(f"  Пунктов: {len(results)}, вариантов: {output['metadata']['total_variants']}")

if __name__ == "__main__":
    main()